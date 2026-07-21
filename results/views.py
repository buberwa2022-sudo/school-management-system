from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.generic import DetailView, ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q, F, Avg, Count, Max, Min, Sum
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.http import FileResponse
import json
from datetime import datetime
import qrcode
from io import BytesIO
import base64

from .models import (
    StudentResult, Examination, Subject, SubjectResult,
    Attendance, BehaviourAssessment, CoCurricularActivity,
    ResultComment, FeeInfo, NextTermInfo, ResultAuditLog,
    GradeScale, DivisionScale, GradingSystem
)
from .utils import (
    calculate_grade, calculate_division, calculate_gpa,
    calculate_ranking, generate_pdf_report, generate_qrcode
)


class StudentResultDetailView(LoginRequiredMixin, DetailView):
    """
    Display comprehensive student result with all sections:
    - School header
    - Student profile
    - Exam info
    - Subject results
    - Academic summary
    - Performance analytics
    - Attendance
    - Behaviour
    - Co-curricular activities
    - Comments
    - Fee info
    - Signatures section
    - QR code
    """
    model = StudentResult
    template_name = 'results/student_result_detail.html'
    context_object_name = 'result'
    pk_url_kwarg = 'pk'
    
    def get_queryset(self):
        # Only authorized users can view results
        user = self.request.user
        queryset = StudentResult.objects.select_related(
            'student', 'examination', 'examination__grading_system'
        ).prefetch_related(
            'subject_results', 'comments', 'activities'
        )
        
        if user.is_staff or user.is_superuser:
            return queryset
        # Students can only view their own results
        if hasattr(user, 'student'):
            return queryset.filter(student=user.student)
        return queryset.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result = self.object
        
        # Fetch all related data
        context['subject_results'] = result.subject_results.select_related('subject').order_by('-average')
        context['comments'] = result.comments.all()
        context['attendance'] = getattr(result, 'attendance', None)
        context['behaviour'] = getattr(result, 'behaviour', None)
        context['activities'] = result.activities.all()
        context['fee_info'] = getattr(result, 'fee_info', None)
        context['next_term_info'] = result.examination.next_term_info
        
        # Performance analysis
        subject_results = context['subject_results']
        if subject_results.exists():
            context['best_subject'] = subject_results.first()
            context['weakest_subject'] = subject_results.last()
            context['performance_trend'] = self._calculate_performance_trend(result)
        
        # Generate QR code
        context['qr_code_base64'] = generate_qrcode(result)
        
        # Grade interpretation
        context['grade_scales'] = result.examination.grading_system.grade_scales.all()
        context['division_scales'] = result.examination.grading_system.division_scales.all()
        
        # Calculate audit log for approvals/changes
        context['audit_logs'] = result.audit_logs.all()[:5]
        
        return context
    
    def _calculate_performance_trend(self, result):
        """Calculate performance trend across previous terms"""
        previous_results = StudentResult.objects.filter(
            student=result.student,
            examination__academic_year__lt=result.examination.academic_year
        ).order_by('-examination__exam_date')
        
        trend = []
        for prev_result in previous_results[:3]:  # Last 3 terms
            trend.append({
                'term': f"{prev_result.examination.get_term_display()} {prev_result.examination.academic_year}",
                'average': prev_result.average_marks,
                'grade': prev_result.overall_grade,
            })
        
        return list(reversed(trend))


class StudentResultListView(LoginRequiredMixin, ListView):
    """
    List all results (admin/staff view)
    """
    model = StudentResult
    template_name = 'results/result_list.html'
    context_object_name = 'results'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = StudentResult.objects.select_related(
            'student', 'examination'
        ).order_by('-examination__exam_date', 'student__user__last_name')
        
        # Staff and superusers can see all results
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            return queryset.none()
        
        # Apply filters
        examination_id = self.request.GET.get('examination')
        if examination_id:
            queryset = queryset.filter(examination_id=examination_id)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(student__user__first_name__icontains=search) |
                Q(student__user__last_name__icontains=search) |
                Q(student__admission_number__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['examinations'] = Examination.objects.all().order_by('-exam_date')
        context['status_choices'] = StudentResult.STATUS_CHOICES
        return context


class ExaminationResultsView(LoginRequiredMixin, TemplateView):
    """
    Dashboard view showing results for a specific examination
    with analytics and statistics
    """
    template_name = 'results/examination_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        examination_id = self.kwargs.get('pk')
        examination = get_object_or_404(Examination, id=examination_id)
        
        results = StudentResult.objects.filter(examination=examination).select_related('student')
        
        # Statistics
        context['examination'] = examination
        context['total_students'] = results.count()
        context['passed_students'] = results.filter(status='pass').count()
        context['failed_students'] = results.filter(status='fail').count()
        context['results'] = results.order_by('-overall_percentage')
        
        # Performance metrics
        stats = results.aggregate(
            avg_marks=Avg('average_marks'),
            highest_marks=Max('average_marks'),
            lowest_marks=Min('average_marks'),
            avg_gpa=Avg('gpa')
        )
        context['stats'] = stats
        
        # Division distribution
        division_stats = results.values('division').annotate(count=Count('id')).order_by('division')
        context['division_stats'] = division_stats
        
        # Subject performance
        subject_stats = SubjectResult.objects.filter(
            student_result__examination=examination
        ).values('subject__name').annotate(
            avg_marks=Avg('average'),
            passed=Count('id', filter=Q(is_passed=True))
        ).order_by('-avg_marks')
        context['subject_stats'] = subject_stats
        
        return context


class ResultPDFExportView(LoginRequiredMixin, DetailView):
    """
    Export student result as PDF
    """
    model = StudentResult
    
    def get(self, request, *args, **kwargs):
        result = self.get_object()
        pdf_file = generate_pdf_report(result)
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{result.student.user.username}_result.pdf"'
        return response
    
    def get_queryset(self):
        user = self.request.user
        queryset = StudentResult.objects.select_related('student')
        
        if user.is_staff or user.is_superuser:
            return queryset
        if hasattr(user, 'student'):
            return queryset.filter(student=user.student)
        return queryset.none()


class ResultExcelExportView(LoginRequiredMixin, DetailView):
    """
    Export examination results as Excel
    """
    model = Examination
    
    def get(self, request, *args, **kwargs):
        examination = self.get_object()
        
        if not (request.user.is_staff or request.user.is_superuser):
            return HttpResponse('Unauthorized', status=403)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('openpyxl not installed', status=500)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Headers
        headers = ['Admission No', 'Name', 'Class', 'Total Marks', 'Average', 'Grade', 'Division', 'Position']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        results = StudentResult.objects.filter(examination=examination).select_related('student').order_by('-overall_percentage')
        for result in results:
            ws.append([
                result.student.admission_number,
                str(result.student.user),
                result.student.current_class,
                result.total_marks,
                f"{result.average_marks:.2f}",
                result.overall_grade,
                result.division,
                result.overall_position,
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{examination.name}_results.xlsx"'
        return response


@staff_member_required
@require_POST
def approve_result(request, pk):
    """
    Approve student result (admin only)
    """
    result = get_object_or_404(StudentResult, pk=pk)
    result.is_approved = True
    result.approved_by = request.user
    result.save()
    
    # Log action
    ResultAuditLog.objects.create(
        student_result=result,
        action='approve',
        user=request.user,
        ip_address=get_client_ip(request)
    )
    
    return JsonResponse({'success': True, 'message': 'Result approved'})


@staff_member_required
@require_POST
def publish_result(request, pk):
    """
    Publish student result (admin only)
    """
    result = get_object_or_404(StudentResult, pk=pk)
    
    if not result.is_approved:
        return JsonResponse({'success': False, 'message': 'Result must be approved first'}, status=400)
    
    result.is_published = True
    result.save()
    
    # Log action
    ResultAuditLog.objects.create(
        student_result=result,
        action='publish',
        user=request.user,
        ip_address=get_client_ip(request)
    )
    
    return JsonResponse({'success': True, 'message': 'Result published'})


@staff_member_required
@require_POST
def lock_result(request, pk):
    """
    Lock result to prevent further modifications
    """
    result = get_object_or_404(StudentResult, pk=pk)
    examination = result.examination
    
    if examination.is_locked:
        return JsonResponse({'success': False, 'message': 'Examination is already locked'}, status=400)
    
    examination.is_locked = True
    examination.save()
    
    return JsonResponse({'success': True, 'message': 'Results locked'})


def get_client_ip(request):
    """
    Get client IP address from request
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


class ResultAnalyticsAPIView(LoginRequiredMixin, TemplateView):
    """
    API endpoint for result analytics (for charts)
    """
    def get(self, request, *args, **kwargs):
        examination_id = request.GET.get('examination')
        if not examination_id:
            return JsonResponse({'error': 'Examination ID required'}, status=400)
        
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        examination = get_object_or_404(Examination, id=examination_id)
        results = StudentResult.objects.filter(examination=examination)
        
        # Subject performance data
        subject_data = SubjectResult.objects.filter(
            student_result__examination=examination
        ).values('subject__name').annotate(
            avg_marks=Avg('average'),
            count=Count('id')
        ).order_by('-avg_marks')
        
        # Grade distribution
        grade_dist = results.values('overall_grade').annotate(count=Count('id')).order_by('overall_grade')
        
        # Division distribution
        division_dist = results.values('division').annotate(count=Count('id')).order_by('division')
        
        return JsonResponse({
            'subject_performance': list(subject_data),
            'grade_distribution': list(grade_dist),
            'division_distribution': list(division_dist),
            'total_results': results.count(),
            'average_marks': results.aggregate(avg=Avg('average_marks'))['avg'],
        })
